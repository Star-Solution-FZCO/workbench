import calendar
from collections.abc import Sequence
from datetime import date, timedelta
from tempfile import NamedTemporaryFile
from typing import Any

import sqlalchemy as sa
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

import wb.models as m
from shared_utils.dateutils import date_range, day_start

__all__ = ('generate_working_time_month_report',)


async def generate_working_time_month_report(
    flt: Any, month: date, session: AsyncSession
) -> Any:
    # pylint: disable=all
    employees_raw = await session.scalars(
        sa.select(m.Employee).filter(flt).order_by(m.Employee.english_name)
    )
    employees: Sequence[m.Employee] = employees_raw.all()

    report_start = month.replace(day=1)
    report_end = month.replace(day=calendar.monthrange(month.year, month.month)[1])
    cur_date = date.today()

    tm_data_raw = await session.scalars(
        sa.select(m.TMRecord)
        .filter(
            m.TMRecord.time >= day_start(report_start),
            m.TMRecord.time < day_start(report_end + timedelta(days=1)),
            m.TMRecord.employee_id.in_([emp.id for emp in employees]),
        )
        .order_by(m.TMRecord.employee_id, m.TMRecord.time)
    )
    tm_data: dict[int, dict[date, list[m.TMRecord]]] = {
        emp.id: {d: [] for d in date_range(report_start, report_end)}
        for emp in employees
    }
    for tm_rec in tm_data_raw.all():
        tm_data[tm_rec.employee_id][tm_rec.time.date()].append(tm_rec)
    for emp in employees:
        transfer_come = False
        for day in date_range(report_start, report_end):
            recs = tm_data[emp.id][day]
            is_current_or_future = day >= cur_date
            if transfer_come:
                recs.insert(
                    0,
                    m.TMRecord(
                        employee_id=emp.id,
                        status=m.TMRecordType.COME,
                        time=day_start(day),
                    ),
                )
                transfer_come = False
            if not recs:
                continue
            last = recs[-1]
            next_day_start = day_start(day + timedelta(days=1))
            if not is_current_or_future and last.status == m.TMRecordType.AWAY:
                recs.append(
                    m.TMRecord(
                        employee_id=emp.id,
                        status=m.TMRecordType.LEAVE,
                        time=next_day_start,
                    )
                )
            elif not is_current_or_future and last.status in (
                m.TMRecordType.AWAKE,
                m.TMRecordType.COME,
            ):
                recs.append(
                    m.TMRecord(
                        employee_id=emp.id,
                        status=m.TMRecordType.LEAVE,
                        time=next_day_start,
                    )
                )
                transfer_come = True

    wb = Workbook()
    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(sheet)
    ws_name = f'{report_start.strftime("%B %Y")}'
    ws = wb.create_sheet(ws_name)
    c = ws['B3']
    ws.freeze_panes = c
    ws.column_dimensions['A'].width = 20
    first_column_index = 2
    weekend_pattern = PatternFill('solid', fgColor='d7e3bb')
    vacation_pattern = PatternFill('solid', fgColor='ccc0da')
    sick_pattern = PatternFill('solid', fgColor='c4bd97')
    is_missed_pattern = PatternFill('solid', fgColor='f9f9f7')
    total_pattern = PatternFill('solid', fgColor='d1d1cd')
    trip_pattern = PatternFill('solid', fgColor='95b3d7')

    for index, day in enumerate(date_range(report_start, report_end)):
        column_index = first_column_index + index * 3
        ws.merge_cells(
            start_row=1,
            start_column=column_index,
            end_row=1,
            end_column=column_index + 2,
        )
        ws.cell(1, column_index).value = day.strftime('%d %b %Y')
        ws.cell(1, column_index).alignment = Alignment(
            horizontal='center', vertical='center'
        )
        ws.cell(2, column_index).value = 'start'
        ws.cell(2, column_index).alignment = Alignment(
            horizontal='center', vertical='center'
        )
        ws.cell(2, column_index + 1).value = 'leave'
        ws.cell(2, column_index + 1).alignment = Alignment(
            horizontal='center', vertical='center'
        )
        ws.cell(2, column_index + 2).value = 'total'
        ws.cell(2, column_index + 2).fill = total_pattern
        ws.cell(2, column_index + 2).alignment = Alignment(
            horizontal='center', vertical='center'
        )

    cell_offset = ((report_end - report_start).days + 1) * 3 + 1
    ws.cell(1, cell_offset + 1).value = 'business trip (hours)'
    ws.cell(1, cell_offset + 1).fill = trip_pattern
    ws.column_dimensions[get_column_letter(cell_offset + 1)].width = 18
    ws.cell(1, cell_offset + 2).value = 'sick days (hours)'
    ws.cell(1, cell_offset + 2).fill = sick_pattern
    ws.column_dimensions[get_column_letter(cell_offset + 2)].width = 20
    ws.cell(1, cell_offset + 3).value = 'vacation (hours)'
    ws.cell(1, cell_offset + 3).fill = vacation_pattern
    ws.column_dimensions[get_column_letter(cell_offset + 3)].width = 15
    ws.cell(1, cell_offset + 4).value = 'in office (hours)'
    ws.column_dimensions[get_column_letter(cell_offset + 4)].width = 23
    ws.cell(1, cell_offset + 5).value = 'total (hours)'
    ws.column_dimensions[get_column_letter(cell_offset + 5)].width = 17
    ws.cell(1, cell_offset + 6).value = 'missed (hours)'
    ws.column_dimensions[get_column_letter(cell_offset + 6)].width = 18
    current_row = 3

    for emp in employees:
        ws.cell(current_row, 1).value = emp.english_name
        workday_hours = 9
        total = timedelta()
        for index, day in enumerate(date_range(report_start, report_end)):
            multiplier = 1 if index == 0 else 3
            if tm_data[emp.id][day]:
                come = tm_data[emp.id][day][0].time.strftime('%H:%M')
                leave = tm_data[emp.id][day][-1].time.strftime('%H:%M')
                total_day = str(
                    tm_data[emp.id][day][-1].time - tm_data[emp.id][day][0].time
                )
                total += tm_data[emp.id][day][-1].time - tm_data[emp.id][day][0].time
            else:
                come = leave = total_day = ''
            ws.cell(current_row, index * multiplier + 2).value = come
            ws.cell(current_row, index * multiplier + 3).value = leave
            ws.cell(current_row, index * multiplier + 4).value = total_day
            ws.cell(current_row, index * multiplier + 4).fill = total_pattern
        ws.cell(current_row, cell_offset + 5).value = str(total)
        current_row += 1

    with NamedTemporaryFile() as tmp_file:
        wb.save(tmp_file.name)
        tmp_file.seek(0)
        return tmp_file.read()
